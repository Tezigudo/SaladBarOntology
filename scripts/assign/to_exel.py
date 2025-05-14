import owlready2
from owlready2 import get_ontology
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os
import rdflib
from rdflib import RDF, RDFS, OWL

def extract_local_name(uri):
    """Extract local name from URI."""
    return str(uri).split('#')[-1] if '#' in str(uri) else str(uri).split('/')[-1]

def export_ontology_to_xlsx_with_swrl(owl_file, output_xlsx):
    # Initialize RDF graph for non-SWRL components (using rdflib)
    g = rdflib.Graph()
    g.parse(owl_file, format='xml')  # Adjust format if needed

    # Load ontology with Owlready2 for SWRL rules
    owlready2.onto_path.append(os.path.dirname(owl_file))  # Add ontology path
    onto = get_ontology(f"file://{owl_file}").load()

    # Initialize Excel workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Sheet 1: Classes and Subclasses (using rdflib)
    ws_classes = wb.create_sheet("Classes")
    ws_classes.append(["Class", "SubClassOf"])
    classes = set()
    for s in g.subjects(RDF.type, OWL.Class):
        if isinstance(s, rdflib.URIRef):
            class_name = extract_local_name(s)
            super_classes = [extract_local_name(o) for o in g.objects(s, RDFS.subClassOf) if isinstance(o, rdflib.URIRef)]
            classes.add((class_name, ", ".join(super_classes)))
    for class_name, super_classes in sorted(classes):
        ws_classes.append([class_name, super_classes])

    # Sheet 2: Individuals (using rdflib)
    ws_individuals = wb.create_sheet("Individuals")
    ws_individuals.append(["Individual", "Type"])
    individuals = set()
    for s in g.subjects(RDF.type):
        if s != OWL.NamedIndividual and isinstance(s, rdflib.URIRef):
            types = [extract_local_name(o) for o in g.objects(s, RDF.type) if o != OWL.NamedIndividual]
            individuals.add((extract_local_name(s), ", ".join(types)))
    for ind_name, types in sorted(individuals):
        ws_individuals.append([ind_name, types])

    # Sheet 3: Properties (using rdflib)
    ws_properties = wb.create_sheet("Properties")
    ws_properties.append(["Property", "Type", "SubPropertyOf", "Domain", "Range"])
    properties = set()
    for s in g.subjects(RDF.type, OWL.ObjectProperty):
        prop_name = extract_local_name(s)
        domains = [extract_local_name(o) for o in g.objects(s, RDFS.domain)]
        ranges = [extract_local_name(o) for o in g.objects(s, RDFS.range)]
        super_props = [extract_local_name(o) for o in g.objects(s, RDFS.subPropertyOf)]
        properties.add((prop_name, "ObjectProperty", ", ".join(super_props), ", ".join(domains), ", ".join(ranges)))
    for s in g.subjects(RDF.type, OWL.DatatypeProperty):
        prop_name = extract_local_name(s)
        domains = [extract_local_name(o) for o in g.objects(s, RDFS.domain)]
        ranges = [extract_local_name(o) for o in g.objects(s, RDFS.range)]
        super_props = [extract_local_name(o) for o in g.objects(s, RDFS.subPropertyOf)]
        properties.add((prop_name, "DatatypeProperty", ", ".join(super_props), ", ".join(domains), ", ".join(ranges)))
    for prop in sorted(properties):
        ws_properties.append(list(prop))

    # Sheet 4: Restrictions (using rdflib)
    ws_restrictions = wb.create_sheet("Restrictions")
    ws_restrictions.append(["Class", "Restriction Type", "Property", "Value"])
    restrictions = set()
    for s in g.subjects(RDF.type, OWL.Restriction):
        on_property = extract_local_name(g.value(s, OWL.onProperty))
        for class_s in g.subjects(RDFS.subClassOf, s):
            class_name = extract_local_name(class_s)
            for pred in [OWL.someValuesFrom, OWL.allValuesFrom, OWL.hasValue]:
                value = g.value(s, pred)
                if value:
                    restrictions.add((class_name, pred.split('#')[-1], on_property, extract_local_name(value) if isinstance(value, rdflib.URIRef) else str(value)))
    for restr in sorted(restrictions):
        ws_restrictions.append(list(restr))

    # Sheet 5: SWRL Rules (using Owlready2)
    ws_rules = wb.create_sheet("Rules")
    ws_rules.append(["Rule Name", "Antecedent", "Consequent"])
    rules = set()
    for rule in onto.rules():
        rule_name = extract_local_name(rule.iri) if hasattr(rule, 'iri') else f"Rule_{id(rule)}"
        # Extract antecedent (body) and consequent (head)
        antecedent = []
        consequent = []
        for atom in rule.body:
            antecedent.append(str(atom))
        for atom in rule.head:
            consequent.append(str(atom))
        rules.add((rule_name, "; ".join(antecedent), "; ".join(consequent)))
    for rule_name, ant, cons in sorted(rules):
        ws_rules.append([rule_name, ant, cons])

    # Sheet 6: All Triples (using rdflib, catch-all)
    ws_triples = wb.create_sheet("All_Triples")
    ws_triples.append(["Subject", "Predicate", "Object"])
    for s, p, o in sorted(g):
        ws_triples.append([extract_local_name(s), extract_local_name(p), extract_local_name(o) if isinstance(o, rdflib.URIRef) else str(o)])

    # Adjust column widths
    for ws in wb:
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 2, 50)

    # Save the workbook
    wb.save(output_xlsx)
    print(f"Ontology with SWRL rules exported to {output_xlsx}")

if __name__ == "__main__":
    # Example usage
    owl_file = "salad_ontology.rdf"  # Replace with your OWL file path
    output_xlsx = "data/salad_ontology.xlsx"
    if os.path.exists(owl_file):
        export_ontology_to_xlsx_with_swrl(owl_file, output_xlsx)
    else:
        print(f"Error: OWL file '{owl_file}' not found.")