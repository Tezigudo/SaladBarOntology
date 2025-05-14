import openpyxl
from rdflib import Graph, Namespace, RDF, RDFS, OWL, Literal, URIRef
import owlready2
from owlready2 import get_ontology, sync_reasoner
import os

def create_uri(namespace, local_name):
    """Create a URI from a namespace and local name."""
    return URIRef(f"{namespace}{local_name}")

def xlsx_to_rdf(xlsx_file, output_owl, namespace="http://example.org/ontology#"):
    # Initialize RDF graph
    g = Graph()
    ns = Namespace(namespace)
    g.bind("owl", OWL)
    g.bind("rdfs", RDFS)
    g.bind("swrl", "http://www.w3.org/2003/11/swrl#")

    # Load Excel file
    wb = openpyxl.load_workbook(xlsx_file)

    # Process Classes sheet
    if "Classes" in wb.sheetnames:
        ws = wb["Classes"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            class_name, super_classes = row
            class_uri = create_uri(ns, class_name)
            g.add((class_uri, RDF.type, OWL.Class))
            if super_classes:
                for super_class in super_classes.split(", "):
                    super_class_uri = create_uri(ns, super_class)
                    g.add((class_uri, RDFS.subClassOf, super_class_uri))

    # Process Individuals sheet
    if "Individuals" in wb.sheetnames:
        ws = wb["Individuals"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            ind_name, types = row
            ind_uri = create_uri(ns, ind_name)
            g.add((ind_uri, RDF.type, OWL.NamedIndividual))
            if types:
                for type_name in types.split(", "):
                    type_uri = create_uri(ns, type_name)
                    g.add((ind_uri, RDF.type, type_uri))

    # Process Properties sheet
    if "Properties" in wb.sheetnames:
        ws = wb["Properties"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            prop_name, prop_type, sub_props, domains, ranges = row
            prop_uri = create_uri(ns, prop_name)
            prop_type_uri = OWL.ObjectProperty if prop_type == "ObjectProperty" else OWL.DatatypeProperty
            g.add((prop_uri, RDF.type, prop_type_uri))
            if sub_props:
                for sub_prop in sub_props.split(", "):
                    sub_prop_uri = create_uri(ns, sub_prop)
                    g.add((prop_uri, RDFS.subPropertyOf, sub_prop_uri))
            if domains:
                for domain in domains.split(", "):
                    domain_uri = create_uri(ns, domain)
                    g.add((prop_uri, RDFS.domain, domain_uri))
            if ranges:
                for range_ in ranges.split(", "):
                    range_uri = create_uri(ns, range_)
                    g.add((prop_uri, RDFS.range, range_uri))

    # Process Restrictions sheet
    if "Restrictions" in wb.sheetnames:
        ws = wb["Restrictions"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            class_name, restr_type, prop_name, value = row
            class_uri = create_uri(ns, class_name)
            prop_uri = create_uri(ns, prop_name)
            restr_uri = URIRef(f"{ns}Restriction_{class_name}_{prop_name}")
            g.add((restr_uri, RDF.type, OWL.Restriction))
            g.add((restr_uri, OWL.onProperty, prop_uri))
            restr_pred = {
                "someValuesFrom": OWL.someValuesFrom,
                "allValuesFrom": OWL.allValuesFrom,
                "hasValue": OWL.hasValue
            }.get(restr_type)
            if restr_pred:
                try:
                    # Try to treat value as a URI; if it fails, treat as literal
                    value_uri = create_uri(ns, value)
                    g.add((restr_uri, restr_pred, value_uri))
                except:
                    g.add((restr_uri, restr_pred, Literal(value)))
            g.add((class_uri, RDFS.subClassOf, restr_uri))

    # Process All_Triples sheet (for additional triples)
    if "All_Triples" in wb.sheetnames:
        ws = wb["All_Triples"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            subj, pred, obj = row
            try:
                subj_uri = create_uri(ns, subj)
                pred_uri = create_uri(ns, pred)
                # Try to treat object as URI; if it fails, treat as literal
                try:
                    obj_uri = create_uri(ns, obj)
                    g.add((subj_uri, pred_uri, obj_uri))
                except:
                    g.add((subj_uri, pred_uri, Literal(obj)))
            except:
                pass  # Skip invalid triples

    # Save RDF graph temporarily (without SWRL rules)
    temp_owl = "temp_ontology.owl"
    g.serialize(temp_owl, format="xml")

    # Process SWRL Rules using Owlready2
    owlready2.onto_path.append(os.getcwd())
    onto = get_ontology(f"file://{temp_owl}").load()
    
    if "Rules" in wb.sheetnames:
        ws = wb["Rules"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rule_name, antecedent, consequent = row
            # Create a new SWRL rule
            with onto:
                rule = owlready2.Imp()
                rule.set_name(rule_name)
                # Parse antecedent and consequent (simplified; assumes atoms are class/property assertions)
                ante_atoms = []
                cons_atoms = []
                for atom in antecedent.split("; "):
                    try:
                        # Example: ClassName(?x) or property(?x, ?y)
                        if "(?" in atom:
                            pred_name = atom[:atom.index("(")]
                            pred = onto.search_one(iri=f"{ns}{pred_name}")
                            if pred:
                                ante_atoms.append(pred)
                    except:
                        pass
                for atom in consequent.split("; "):
                    try:
                        if "(?" in atom:
                            pred_name = atom[:atom.index("(")]
                            pred = onto.search_one(iri=f"{ns}{pred_name}")
                            if pred:
                                cons_atoms.append(pred)
                    except:
                        pass
                # Assign simplified atoms (this is a basic implementation)
                if ante_atoms and cons_atoms:
                    rule.set_as_rule(f"{antecedent} -> {consequent}")

    # Save the final ontology with SWRL rules
    onto.save(file=output_owl, format="rdfxml")
    print(f"Ontology reconstructed and saved to {output_owl}")

    # Clean up temporary file
    if os.path.exists(temp_owl):
        os.remove(temp_owl)

if __name__ == "__main__":
    # Example usage
    xlsx_file = "ontology_output_with_swrl.xlsx"  # Input XLSX file
    output_owl = "reconstructed_ontology.owl"  # Output OWL file
    if os.path.exists(xlsx_file):
        xlsx_to_rdf(xlsx_file, output_owl)
    else:
        print(f"Error: XLSX file '{xlsx_file}' not found.")